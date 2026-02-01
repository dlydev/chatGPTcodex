#!/usr/bin/env python3
"""Menu-driven bid tools for creating bid folders and updating the bid list workbook."""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import shutil
from dataclasses import dataclass
from pathlib import Path

import openpyxl


HEADER_DEFAULTS = [
    "Bid Folder",
    "Bid Number",
    "Estimator",
    "Bid Due Date",
    "Customer/GC",
    "Bid Name",
    "Proposal Date",
    "Proposal Amount",
    "Bid Status",
]


def sanitize_name(value: str | None) -> str:
    if value is None:
        return ""
    value = re.sub(r"[\\/:*?\"<>|]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def read_yes_no_default_no(prompt: str) -> bool:
    while True:
        raw = input(f"{prompt} (Y/N) [N]").strip()
        if not raw:
            return False
        if re.match(r"^(y|yes)$", raw, re.IGNORECASE):
            return True
        if re.match(r"^(n|no)$", raw, re.IGNORECASE):
            return False
        print("Please enter Y or N (or press Enter for N).")


def read_non_empty(prompt: str) -> str:
    while True:
        value = sanitize_name(input(prompt))
        if value.strip():
            return value
        print("Value is required.")


def assert_paths(bid_root: Path, template_root: Path) -> None:
    if not bid_root.exists():
        raise FileNotFoundError(f"BidRoot not found: {bid_root}")
    if not template_root.exists():
        raise FileNotFoundError(f"TemplateRoot not found: {template_root}")


def get_bid_folders(bid_root: Path) -> list[Path]:
    return sorted([p for p in bid_root.iterdir() if p.is_dir()], key=lambda p: p.name)


def get_next_bid_number(bid_root: Path) -> int:
    max_value = 0
    for folder in get_bid_folders(bid_root):
        match = re.match(r"^\s*(\d+)\b", folder.name)
        if match:
            number = int(match.group(1))
            if number > max_value:
                max_value = number
    if max_value == 0:
        raise ValueError(f"No existing bid-number folders found in: {bid_root}")
    return max_value + 1


def normalize_bid_date(raw_date: str) -> str:
    match = re.match(r"^(0?[1-9]|1[0-2])-(0?[1-9]|[12]\d|3[01])$", raw_date)
    if not match:
        raise ValueError(
            "Bid Date must be in MM-DD format (ex: 12-5 or 12-05). "
            f"You entered: {raw_date}"
        )
    month = int(match.group(1))
    day = int(match.group(2))
    return f"{month:02d}-{day:02d}"


def build_bid_folder_name(
    bid_number: int, initials: str, bid_date: str, customer: str, bid_name: str
) -> str:
    name = f"{bid_number} - {initials} - {bid_date} - {customer} - {bid_name}"
    return sanitize_name(name)


@dataclass
class BidFolderInfo:
    bid_number: str
    initials: str
    bid_date: str
    customer: str
    bid_name: str
    folder: str


def parse_bid_folder_name(folder_name: str) -> BidFolderInfo | None:
    parts = re.split(r"\s-\s", folder_name, maxsplit=4)
    if len(parts) < 5:
        return None
    return BidFolderInfo(
        bid_number=parts[0].strip(),
        initials=parts[1].strip(),
        bid_date=parts[2].strip(),
        customer=parts[3].strip(),
        bid_name=parts[4].strip(),
        folder=folder_name,
    )


def get_pending_save_path(path: Path) -> Path:
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    return path.with_name(f"{path.stem} - Pending Update {stamp}{path.suffix}")


@dataclass
class ExcelContext:
    workbook: openpyxl.Workbook
    worksheet: openpyxl.worksheet.worksheet.Worksheet
    read_only: bool
    pending_save_path: Path | None
    workbook_path: Path


def new_excel_context(path: Path, worksheet_name: str) -> ExcelContext:
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    read_only = False
    pending_save_path: Path | None = None

    try:
        workbook = openpyxl.load_workbook(path)
    except PermissionError:
        read_only = True
        pending_save_path = get_pending_save_path(path)
        shutil.copy2(path, pending_save_path)
        workbook = openpyxl.load_workbook(pending_save_path)

    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.create_sheet(worksheet_name)

    return ExcelContext(
        workbook=workbook,
        worksheet=worksheet,
        read_only=read_only,
        pending_save_path=pending_save_path,
        workbook_path=path,
    )


def close_excel_context(context: ExcelContext) -> None:
    if context.read_only and context.pending_save_path is not None:
        context.workbook.save(context.pending_save_path)
    else:
        context.workbook.save(context.workbook_path)


def ensure_headers(worksheet) -> dict[str, int]:
    legacy_header_map = {
        "Folder Name": "Bid Folder",
        "Bid#": "Bid Number",
        "GC/Owner": "Customer/GC",
        "Description": "Bid Name",
        "Due Date": "Bid Due Date",
        "Status": "Bid Status",
    }

    headers: dict[str, int] = {}
    has_any = False
    for col in range(1, 31):
        value = worksheet.cell(row=1, column=col).value
        if value:
            canonical = legacy_header_map.get(str(value), str(value))
            if canonical != value:
                worksheet.cell(row=1, column=col).value = canonical
            headers.setdefault(canonical, col)
            has_any = True

    if not has_any:
        for idx, header in enumerate(HEADER_DEFAULTS, start=1):
            worksheet.cell(row=1, column=idx).value = header
            headers[header] = idx
        return headers

    needs_reorder = False
    for idx, header in enumerate(HEADER_DEFAULTS, start=1):
        current = worksheet.cell(row=1, column=idx).value
        if current != header:
            needs_reorder = True
            break

    if needs_reorder:
        for idx, header in enumerate(HEADER_DEFAULTS, start=1):
            worksheet.cell(row=1, column=idx).value = header

    for idx, header in enumerate(HEADER_DEFAULTS, start=1):
        headers[header] = idx

    return headers


def get_last_row(worksheet) -> int:
    return worksheet.max_row or 1


def get_cell_text(worksheet, row: int, col: int) -> str:
    value = worksheet.cell(row=row, column=col).value
    if value is None:
        return ""
    return str(value).strip()


def get_row_index_by_value(worksheet, col: int, target: str) -> int | None:
    for row in range(2, get_last_row(worksheet) + 1):
        value = worksheet.cell(row=row, column=col).value
        if value is None:
            continue
        if str(value).strip() == target:
            return row
    return None


def write_row(worksheet, headers: dict[str, int], row: int, bid_info: BidFolderInfo) -> None:
    worksheet.cell(row=row, column=headers["Bid Folder"]).value = bid_info.folder
    worksheet.cell(row=row, column=headers["Bid Number"]).value = bid_info.bid_number
    worksheet.cell(row=row, column=headers["Estimator"]).value = bid_info.initials
    worksheet.cell(row=row, column=headers["Bid Due Date"]).value = bid_info.bid_date
    worksheet.cell(row=row, column=headers["Customer/GC"]).value = bid_info.customer
    worksheet.cell(row=row, column=headers["Bid Name"]).value = bid_info.bid_name


def sync_bid_workbook(bid_root: Path, workbook_path: Path, worksheet_name: str) -> None:
    context = new_excel_context(workbook_path, worksheet_name)
    try:
        worksheet = context.worksheet
        headers = ensure_headers(worksheet)
        last_row = get_last_row(worksheet)

        for folder in get_bid_folders(bid_root):
            info = parse_bid_folder_name(folder.name)
            if info is None:
                continue
            row = get_row_index_by_value(worksheet, headers["Bid Number"], info.bid_number)
            if row is None:
                row = get_row_index_by_value(worksheet, headers["Bid Folder"], info.folder)
            if row is None:
                last_row += 1
                row = last_row
            write_row(worksheet, headers, row, info)
    finally:
        close_excel_context(context)

    if context.read_only and context.pending_save_path is not None:
        print("Workbook is open by another user; saved updates to:")
        print(context.pending_save_path)
    else:
        print("Workbook updated with current bid folders.")


def update_bid_status(workbook_path: Path, worksheet_name: str) -> None:
    bid_number = read_non_empty("Enter bid number to update")
    status = input("Status (leave blank to keep current)").strip()
    award = ""
    proposal_amount = ""
    proposal_date = ""

    context = new_excel_context(workbook_path, worksheet_name)
    try:
        worksheet = context.worksheet
        headers = ensure_headers(worksheet)
        row = get_row_index_by_value(worksheet, headers["Bid Number"], bid_number)
        if row is None:
            raise ValueError(f"Bid number not found in workbook: {bid_number}")

        if "Proposal Amount" in headers:
            current_amount = get_cell_text(worksheet, row, headers["Proposal Amount"])
            if current_amount:
                if read_yes_no_default_no(
                    f"Proposal Amount is '{current_amount}'. Update it?"
                ):
                    proposal_amount = input("Proposal Amount").strip()
            else:
                proposal_amount = input("Proposal Amount (leave blank to skip)").strip()

        if "Proposal Date" in headers:
            current_date = get_cell_text(worksheet, row, headers["Proposal Date"])
            if current_date:
                if read_yes_no_default_no(
                    f"Proposal Date is '{current_date}'. Update it?"
                ):
                    proposal_date = input("Proposal Date (MM-DD or date string)").strip()
            else:
                proposal_date = input("Proposal Date (leave blank to skip)").strip()

        if "Award" in headers:
            award = input("Award (leave blank to keep current)").strip()

        if status and "Bid Status" in headers:
            worksheet.cell(row=row, column=headers["Bid Status"]).value = status

        if proposal_amount and "Proposal Amount" in headers:
            worksheet.cell(row=row, column=headers["Proposal Amount"]).value = proposal_amount

        if proposal_date and "Proposal Date" in headers:
            worksheet.cell(row=row, column=headers["Proposal Date"]).value = proposal_date

        if award and "Award" in headers:
            worksheet.cell(row=row, column=headers["Award"]).value = award
    finally:
        close_excel_context(context)

    if context.read_only and context.pending_save_path is not None:
        print("Workbook is open by another user; saved updates to:")
        print(context.pending_save_path)
    else:
        print("Workbook status updated.")


def new_bid_folder(bid_root: Path, template_root: Path, workbook_path: Path, worksheet_name: str) -> None:
    assert_paths(bid_root, template_root)

    sync_bid_workbook(bid_root, workbook_path, worksheet_name)

    initials = read_non_empty("Estimator initials (ex: MD)")
    bid_date_raw = read_non_empty("Bid Due Date (MM-DD, ex: 12-5)")
    customer = read_non_empty("Customer/GC")
    bid_name = read_non_empty("Bid Name")

    bid_date = normalize_bid_date(bid_date_raw)
    new_number = get_next_bid_number(bid_root)

    new_folder_name = build_bid_folder_name(new_number, initials, bid_date, customer, bid_name)
    dest = bid_root / new_folder_name
    if dest.exists():
        raise FileExistsError(f"Destination already exists: {dest}")

    dest.mkdir(parents=True)

    if read_yes_no_default_no("Copy subfolder structure from the template?"):
        for item in template_root.iterdir():
            if item.name.lower() == "thumbs.db":
                continue
            target = dest / item.name
            if item.is_dir():
                shutil.copytree(item, target, dirs_exist_ok=True)
            else:
                shutil.copy2(item, target)

        for root, _, files in os.walk(dest):
            for filename in files:
                if filename.lower() == "thumbs.db":
                    try:
                        os.remove(Path(root) / filename)
                    except OSError:
                        continue

    print("\nCreated new bid folder:")
    print(dest)
    print("")

    if read_yes_no_default_no("Update the bid list workbook now?"):
        sync_bid_workbook(bid_root, workbook_path, worksheet_name)

    if os.name == "nt":
        os.startfile(dest)  # noqa: S606
    else:
        print(f"Open folder: {dest}")


def show_menu() -> None:
    print("\nBid Tools")
    print("1) Create new bid folder")
    print("2) Sync bid list workbook with folders")
    print("3) Update bid status in workbook")
    print("4) Exit")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Bid tools")
    parser.add_argument("--bid-root", default=r"S:\\Bid Documents 2026")
    parser.add_argument(
        "--template-root",
        default=r"S:\\Bid Documents 2026\\26000 Proposal Templates\\15 - Folder Structure",
    )
    parser.add_argument(
        "--workbook-path",
        default=r"S:\\Bid Documents 2026\\26000 Proposal Templates\\Bid List.xlsx",
    )
    parser.add_argument("--worksheet-name", default="Bid List")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    bid_root = Path(args.bid_root)
    template_root = Path(args.template_root)
    workbook_path = Path(args.workbook_path)
    worksheet_name = args.worksheet_name

    while True:
        show_menu()
        choice = input("Choose an option (1-4)").strip()
        if choice == "1":
            new_bid_folder(bid_root, template_root, workbook_path, worksheet_name)
        elif choice == "2":
            sync_bid_workbook(bid_root, workbook_path, worksheet_name)
        elif choice == "3":
            update_bid_status(workbook_path, worksheet_name)
        elif choice == "4":
            break
        else:
            print("Invalid option. Choose 1-4.")


if __name__ == "__main__":
    main()
